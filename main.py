import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
from openpyxl.utils import column_index_from_string

def read_credentials(file_path):
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        header = {cell.value: column_index_from_string(cell.column_letter) for cell in sheet[1]}

        if 'Email' not in header or 'Password' not in header:
            raise ValueError("Missing 'Email' or 'Password' column in Excel file header")

        email_value = sheet.cell(row=2, column=header['Email']).value
        password_value = sheet.cell(row=2, column=header['Password']).value

        if not email_value or not password_value:
            raise ValueError("Email or Password is missing in the second row of the Excel file")

        return email_value, password_value

    except FileNotFoundError:
        raise FileNotFoundError(f"File not found at path: {file_path}")

    except KeyError as e:
        raise KeyError(f"KeyError: {e}. Make sure 'Email' and 'Password' columns are present in the Excel file header")

    except Exception as e:
        raise e

def main():
    service = Service(executable_path="chromedriver.exe")
    driver = webdriver.Chrome(service=service)

    driver.get("https://google.com")

    input_element = driver.find_element(By.NAME, "q")
    input_element.clear()
    input_element.send_keys("LinkedIn" + Keys.ENTER)

    WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.PARTIAL_LINK_TEXT, "LinkedIn"))
    )

    link = driver.find_element(By.PARTIAL_LINK_TEXT, "LinkedIn")
    link.click()

    WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.LINK_TEXT, "Sign in"))
    )

    sign_in_link = driver.find_element(By.LINK_TEXT, "Sign in")
    sign_in_link.click()

    WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.NAME, "session_key"))
    )

    email_value, password_value = read_credentials(r'D:\vs codes\credentials.xlsx')

    print(driver.page_source)

    email_field = driver.find_element(By.NAME, "session_key")
    password_field = driver.find_element(By.NAME, "session_password")

    
    email_field.clear()
    password_field.clear()

    email_field.send_keys(email_value)
    password_field.send_keys(password_value)

    login_button = driver.find_element(By.XPATH, "//button[@type='submit']")
    login_button.click()

    # To check whether the login is a success or not
    WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.ID, "some_element_post_login"))  # Replace with an actual element
    )

    time.sleep(30)
    driver.quit()

if __name__ == "__main__":
    main()
